import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

url = "https://data.90tiyu.com/zh-cn/league/europe/england/englandpremierleague/stat/goaltime/"

response = requests.get(url)
content = response.content

soup = BeautifulSoup(content, 'html.parser')

table = soup.find('table')

# 使用pandas的read_html函数直接读取网页中的表格数据
dfs = pd.read_html(str(table))

# 假设表格数据在第一个DataFrame中
df = dfs[0]

# 将'阿森纳'替换为'Arsenal'
df = df.replace('阿森纳', 'Arsenal')
df = df.replace('利物浦', 'Liverpool')
df = df.replace('富勒姆', 'Fulham')
df = df.replace('南安普顿', 'Southampton')
df = df.replace('托特纳姆热刺', 'Tottenham Hotspur')
df = df.replace('纽卡斯尔联', 'Newcastle')
df = df.replace('狼队', 'Wolverhampton Wanderers')
df = df.replace('伯恩茅斯', 'AFC Bournemouth')
df = df.replace('利兹联', 'Leeds United')
df = df.replace('切尔西', 'Chelsea FC')
df = df.replace('曼彻斯特联', 'Manchester United')
df = df.replace('莱斯特城', 'Leicester City')
df = df.replace('布伦特福德', 'Brentford')
df = df.replace('布莱顿', 'Brighton & Hove Albion')
df = df.replace('曼彻斯特城', 'Manchester City')
df = df.replace('埃弗顿', 'Everton')
df = df.replace('阿斯顿维拉', 'Aston Villa')
df = df.replace('诺丁汉森林', 'Nottingham Forest')
df = df.replace('水晶宫', 'Crystal Palace')
df = df.replace('西汉姆联', 'West Ham United')

# 设置Excel文件名和工作表名
filename = 'table_data.xlsx'
sheetname = '数据'

# 创建一个新的Excel文件
wb = Workbook()
ws = wb.active
ws.title = sheetname

# 设置表格数据
for row in dataframe_to_rows(df, index=False, header=True):
    ws.append(row)

# 设置单元格对齐方式
for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
# 设置表头合并单元格
ws.merge_cells('A1:K1')
ws.merge_cells('A2:A3')
ws.merge_cells('B2:K2')

# 自动调整列宽
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        try:
            if cell.value:
                column_letter = cell.column_letter
                column_width = len(str(cell.value)) + 2
                ws.column_dimensions[column_letter].width = column_width
        except:
            pass

# 保存Excel文件
wb.save(filename)
